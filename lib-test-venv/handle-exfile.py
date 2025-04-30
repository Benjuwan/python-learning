import openpyxl

# ## ワークブックを作成
# ```
# ワークブック = openpyxl.Workbook() # Workbook は大文字なので注意
# ```
# ワークブックはExcelのブックに相当し、自動的にシートも生成される。
sample_workbook = openpyxl.Workbook()

# ## ワークブックの保存（新規作成）
# ```
# ワークブック.save('ファイル名.xlsx')
# ```
# 同名ファイルが存在する場合は上書き保存され、無い場合は新規作成となる。
# sample_workbook.save("../anothers/sample_workbook.xlsx")


# ## シートの操作
# （各シートの）セルを読み書きするには、操作対象とするブック（オブジェクト）を選択（アクティブに）する必要がある。
# ```
# ワークシート = ワークブック.active
# ```
sample_worksheet = sample_workbook.active

# ### ワークシート[セル位置]で操作
# 読み込み：ワークシート[セル位置].value
# 書き込み：ワークシート[セル位置] = 値
sample_worksheet["A1"] = "商品名"
sample_worksheet["B1"] = "価格"
# print(sample_worksheet["A1"].value, sample_worksheet["B1"].value)

sample_worksheet["A2"] = "hat"
sample_worksheet["B2"] = "2000"

# 上書き保存
# sample_workbook.save("../anothers/sample_workbook.xlsx")

# 既存Excelファイルの読み込み
sample_workbook_load = openpyxl.load_workbook("../anothers/sample_workbook_load.xlsx")
sample_worksheet_load = sample_workbook_load.active
# print(sample_worksheet_load["A2"].value, sample_worksheet_load["B2"].value)

catalogs = [("hat", 2000), ("shirt", 1000), ("socks", 500)]
sample_worksheet_load["A1"] = "商品名"
sample_worksheet_load["B1"] = "価格"
for i, (name, price) in enumerate(catalogs, 2):
    print(f"A{i}：{name} | B{i}：{price}")
    # 商品名と価格表示を先頭にするため、enumerate は 2 から開始
    sample_worksheet_load[f"A{i}"] = name
    sample_worksheet_load[f"B{i}"] = price

# 既存Excelファイルの上書き保存
sample_workbook_load.save("../anothers/sample_workbook_load.xlsx")

# ## セルを 一行分（iter_rows）/一列分（iter_cols） ずつ取得（5つの省略可能なキーワード引数を受け取る）
# for 文のイテラブルに指定することで各行/列のセルを操作可能。
# ```
# ワークシート.iter_rows(
# min_row=最小の行番号{1〜},
# max_row=最大の行番号{1〜},
# min_col=最小の列番号{1〜},
# max_col=最大の列番号{1〜},
# values_only=Trueを指定するとセルの値だけを返す
# )```

for item in sample_worksheet_load.iter_rows(values_only=True):
    print(item)
    # ('商品名', '価格')
    # ('hat', 2000)
    # ('shirt', 1000)
    # ('socks', 500)

for item in sample_worksheet_load.iter_cols(values_only=True):
    print(item)
    # ('商品名', 'hat', 'shirt', 'socks')
    # ('価格', 2000, 1000, 500)


def update_sheet(targetSampleSheetPass: str) -> None:
    sample_worksheet_load["A1"] = "商品名"
    sample_worksheet_load["B1"] = "価格"
    sum_price = 0
    for i, (_, price) in enumerate(catalogs, 2):
        if price > 0:
            sum_price += price
        print(sum_price)
    sample_worksheet_load[f"A{i + 2}"] = "合計"
    sample_worksheet_load[f"B{i + 2}"] = sum_price

    sample_workbook_load.save(targetSampleSheetPass)


update_sheet("../anothers/sample_workbook_load.xlsx")
